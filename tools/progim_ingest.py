"""Ingest an uploaded Progim workbook into the engine's data files.

The naive full re-extract regresses the curated rules — the workbook's demo
worker reads 0 for components it lacks, so extract_rules loses rules and
mis-derives rates. So ingest is a SAFE MERGE:

  * lookups.json  — taken fully from the upload (DARGA/vetek/caps are what
                    actually change between Progim versions, and extract cleanly).
  * component_rules.json — the curated rules are kept as-is for rates, type,
                    grade_split and the manual special rules; only each percent
                    rule's BASE_CODES are refreshed from the workbook's SACHAR
                    formulas (this is what has been changing — 1233, 1082, the
                    4169 removal — and it reads reliably from the formula).

`ingest(xlsm_path, current_rules)` returns (lookups, rules, summary) without
touching disk, so the caller decides where to persist.
"""
import re
from openpyxl.utils import get_column_letter, column_index_from_string

from tools import extract_lookups


SACHAR_CODE_ROW, SACHAR_FORMULA_ROW = 9, 11
SACHAR_FIRST, SACHAR_LAST = "AA", "DV"


def _sachar_col_codes(ws):
    """Map each SACHAR data column-letter -> set of pay codes it carries.

    Column headers may hold an alias pair, e.g. '1037 / 4544'; both count.
    """
    col_codes = {}
    for c in range(column_index_from_string(SACHAR_FIRST),
                   column_index_from_string(SACHAR_LAST) + 1):
        raw = ws.cell(row=SACHAR_CODE_ROW, column=c).value
        if raw is None:
            continue
        codes = set()
        for part in str(raw).replace(" ", "").split("/"):
            if part.isdigit():
                codes.add(int(part))
        if codes:
            col_codes[get_column_letter(c)] = codes
    return col_codes


def _formula_base_codes(ws, col_codes, primary_codes):
    """Base codes referenced by the row-11 formula of the column that owns any
    of `primary_codes`. Returns None when the rule has no SACHAR column/formula.
    """
    target = set(primary_codes)
    for L, codes in col_codes.items():
        if codes & target:
            f = ws.cell(row=SACHAR_FORMULA_ROW,
                        column=column_index_from_string(L)).value
            if not isinstance(f, str):
                return None
            base = set()
            for ref in re.findall(r"([A-Z]{1,2})11", f):
                base |= col_codes.get(ref, set())
            base -= target  # a rule never sits in its own base
            return base
    return None


def ingest(xlsm_path, current_rules):
    """Merge an uploaded Progim into (lookups, rules). Pure function.

    current_rules: the live component_rules.json dict (str keys). Returned rules
    keep everything curated and only refresh percent rules' base_codes from the
    workbook — never removing a curated non-formula code blindly, and never
    touching rates/type/grade_split/manual rules.
    """
    lookups = extract_lookups.extract(xlsm_path)

    import openpyxl
    wb = openpyxl.load_workbook(xlsm_path, data_only=False)
    changes = []
    if "SACHAR" in wb.sheetnames:
        ws = wb["SACHAR"]
        col_codes = _sachar_col_codes(ws)
        for key, rule in current_rules.items():
            if rule.get("type") != "percent":
                continue
            primaries = rule.get("codes", [])
            formula_base = _formula_base_codes(ws, col_codes, primaries)
            if formula_base is None:
                continue  # rule not represented in SACHAR — leave it untouched
            old = set(rule.get("base_codes", []))
            # ADD-ONLY: never drop a curated code (the curation carries alias
            # pairs like 1037/4544 and codes the demo-worker formula can't
            # express, so a formula-driven removal would regress files that use
            # them). We only ADD base codes the workbook newly references — this
            # captures a new pensionable code entering the bases (1233, 1082,
            # …) automatically, while removals stay a deliberate manual edit.
            new = old | formula_base
            if new != old:
                added = sorted(new - old)
                changes.append({"code": key, "name": rule.get("name", ""),
                                "added": added, "removed": []})
                rule["base_codes"] = sorted(new)

    summary = {
        "grades": len(lookups.get("darga", {})),
        "tracks": len(lookups.get("vetek", {})),
        "rules": len(current_rules),
        "base_changes": changes,
    }
    return lookups, current_rules, summary
