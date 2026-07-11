"""
validate_all.py — run the simulator on EVERY employee in a גולמי file and
compare the computed result against each employee's real payslip.

For each employee the engine recomputes the base component(s) from the state
pay tables (grade × seniority-multiplier × job%), takes the remaining
components from the slip, and classifies the slip:

    תקין (valid)          — computed result matches the real slip (≤ ₪1)
    שגוי (invalid)        — active slip that does not match → needs review
    ללא שכר בסיס פעיל     — pensioner / inactive (base sum ≈ 0)
    רטרו / רב-תקופתי      — more than one pay period on the slip

Outputs (written next to the input file, or to --out DIR):
    validation_per_employee.xlsx — one row per employee: computed vs. slip,
                                   diff, status, flagged components
    VALIDATION_REPORT.md         — Hebrew summary: totals, per-ministry and
                                   per-grade breakdowns, every invalid slip

Usage:
    python tools/validate_all.py golmi.xlsx [--out reports/]
"""

import argparse
import sys
import time
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import main as engine

STATUS_HE = {
    "valid": "תקין",
    "invalid": "שגוי",
    "no_base": "ללא שכר בסיס פעיל",
    "multi_period": "רטרו / רב-תקופתי",
}

SUMMARY_COLS = [
    ("worker_id", "מסד עובד"),
    ("ministry_code", "קוד משרד"),
    ("ministry_name", "שם משרד"),
    ("darga_label", "דרגה (מזוהה)"),
    ("vatek", "ותק"),
    ("job_pct", "חלקיות"),
    ("grade_base", "שכר יסוד (טבלה)"),
    ("vatek_mult", "מקדם ותק"),
    ("base_slip", "בסיס בתלוש"),
    ("base_calc", "בסיס מחושב"),
    ("base_diff", "הפרש בסיס"),
    ("total_expected", "סכום כולל בתלוש"),
    ("total_calculated", "סכום כולל מחושב"),
    ("total_diff", "הפרש כולל"),
    ("status_he", "סטטוס"),
    ("flagged_components", "רכיבים חריגים (חוקה)"),
    ("errors", "הערות"),
]


def run_validation(golmi_path: str):
    """Run the engine on every employee; return the per-employee DataFrame."""
    lookups = engine.get_lookups()
    workers_raw = engine.load_golmi(golmi_path)
    entries = engine.run_engine_full(workers_raw, lookups)

    rows = []
    for entry in entries:
        r, flags = entry["result"], entry["comp_flags"]
        base_slip = sum((c.expected or 0.0) for c in r.components
                        if c.code in engine.BASE_CODES)
        base_calc = sum(c.amount for c in r.components if c.calculated)
        rows.append({
            "worker_id": r.worker_id,
            "ministry_code": r.ministry_code,
            "ministry_name": r.ministry_name,
            "darga_label": r.darga_label,
            "vatek": r.vatek_calculated,
            "job_pct": r.job_pct,
            "grade_base": r.grade_base,
            "vatek_mult": round(r.vatek_multiplier, 6) if r.vatek_multiplier else None,
            "base_slip": round(base_slip, 2),
            "base_calc": round(base_calc, 2) if base_calc else None,
            "base_diff": round(base_calc - base_slip, 2) if base_calc else None,
            "total_expected": r.expected_total,
            "total_calculated": r.total,
            "total_diff": r.total_diff,
            "status": r.status,
            "status_he": STATUS_HE[r.status],
            "flagged_components": "; ".join(
                f"{code} ({chk['name']}): {chk['slip']} במקום {chk['expected']}"
                for code, chk in sorted(flags.items())),
            "errors": "; ".join(r.errors),
        })
    return pd.DataFrame(rows)


def write_excel(df: pd.DataFrame, path: Path):
    """Per-employee sheet, RTL, invalid rows highlighted."""
    out = df[[c for c, _ in SUMMARY_COLS]].rename(columns=dict(SUMMARY_COLS))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="אימות פר עובד", index=False)
        ws = writer.sheets["אימות פר עובד"]
        ws.sheet_view.rightToLeft = True
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        yellow = PatternFill(start_color="FFFFF2AB", end_color="FFFFF2AB",
                             fill_type="solid")
        status_col = [c for c, _ in SUMMARY_COLS].index("status_he") + 1
        for row in ws.iter_rows(min_row=2):
            if row[status_col - 1].value == STATUS_HE["invalid"]:
                for cell in row:
                    cell.fill = yellow
        for i, (key, _) in enumerate(SUMMARY_COLS, start=1):
            width = 22 if key in ("ministry_name", "flagged_components", "errors") else 13
            ws.column_dimensions[get_column_letter(i)].width = width


def write_markdown(df: pd.DataFrame, golmi_path: str, elapsed: float, path: Path):
    c = Counter(df["status"])
    active = c["valid"] + c["invalid"]
    acc = c["valid"] / active * 100 if active else 0.0

    lines = []
    lines.append("# דוח אימות — סימולטור מול תלושי אמת")
    lines.append("")
    lines.append(f"קובץ נבדק: `{Path(golmi_path).name}` · "
                 f"עובדים: **{len(df):,}** · זמן ריצה: {elapsed:.1f} שניות")
    lines.append("")
    lines.append("הסימולטור הורץ על **כל עובד** בקובץ: רכיב הבסיס חושב מטבלאות "
                 "השכר (דרגה × מקדם ותק × חלקיות), הושווה לתלוש בפועל, ורכיבי "
                 "התוספות אומתו לפי חוקת ה-Progim.")
    lines.append("")
    lines.append("## תוצאה כוללת")
    lines.append("")
    lines.append("| סטטוס | עובדים | % מהקובץ |")
    lines.append("|---|---:|---:|")
    for status in ("valid", "invalid", "no_base", "multi_period"):
        n = c.get(status, 0)
        lines.append(f"| {STATUS_HE[status]} | {n:,} | {n / len(df) * 100:.1f}% |")
    lines.append("")
    lines.append(f"**תקינות על התלושים הפעילים: {acc:.2f}%** "
                 f"({c['valid']:,} מתוך {active:,}).")
    lines.append("")

    lines.append("## פילוח לפי משרד")
    lines.append("")
    lines.append("| משרד | עובדים | תקין | שגוי | % תקינות (פעילים) |")
    lines.append("|---|---:|---:|---:|---:|")
    grp = df.groupby("ministry_name")
    stats = pd.DataFrame({
        "workers": grp.size(),
        "valid": grp.apply(lambda g: (g["status"] == "valid").sum(), include_groups=False),
        "invalid": grp.apply(lambda g: (g["status"] == "invalid").sum(), include_groups=False),
    }).sort_values("workers", ascending=False)
    for name, r in stats.iterrows():
        act = r["valid"] + r["invalid"]
        pct = f"{r['valid'] / act * 100:.1f}%" if act else "—"
        lines.append(f"| {name} | {r['workers']:,} | {r['valid']:,} | "
                     f"{r['invalid']:,} | {pct} |")
    lines.append("")

    lines.append("## פילוח לפי דרגה (מזוהה)")
    lines.append("")
    lines.append("| דרגה | עובדים | תקין | שגוי |")
    lines.append("|---|---:|---:|---:|")
    grp = df[df["status"].isin(["valid", "invalid"])].groupby("darga_label")
    for label, g in sorted(grp, key=lambda kv: str(kv[0])):
        nv = (g["status"] == "valid").sum()
        lines.append(f"| {label} | {len(g):,} | {nv:,} | {len(g) - nv:,} |")
    lines.append("")

    inv = df[df["status"] == "invalid"].sort_values("base_diff",
                                                    key=lambda s: s.abs(),
                                                    ascending=False)
    lines.append(f"## תלושים שגויים — לבדיקה ({len(inv)})")
    lines.append("")
    if len(inv):
        lines.append("| מסד | משרד | דרגה | ותק | חלקיות | בסיס בתלוש | בסיס מחושב | הפרש |")
        lines.append("|---|---|---|---:|---:|---:|---:|---:|")
        for _, r in inv.iterrows():
            lines.append(
                f"| {r['worker_id']} | {r['ministry_name']} | {r['darga_label']} "
                f"| {r['vatek']} | {r['job_pct']} | {r['base_slip']:,.2f} "
                f"| {r['base_calc']:,.2f} | {r['base_diff']:,.2f} |")
        lines.append("")
        lines.append("הפערים האופייניים: כניסה/עזיבה באמצע חודש (בסיס חלקי), "
                     "שינוי דרגה באמצע חודש, תשלום רטרו שנבלע בשורת הבסיס, "
                     "וותק מעבר לתקרת הטבלה.")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main_cli():
    ap = argparse.ArgumentParser(description="Validate the simulator against every employee's real payslip.")
    ap.add_argument("golmi", help="path to the גולמי .xlsx file")
    ap.add_argument("--out", default="reports", help="output directory (default: reports/)")
    args = ap.parse_args()

    t0 = time.time()
    df = run_validation(args.golmi)
    elapsed = time.time() - t0

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "validation_per_employee.xlsx"
    md_path = out_dir / "VALIDATION_REPORT.md"
    write_excel(df, xlsx_path)
    write_markdown(df, args.golmi, elapsed, md_path)

    c = Counter(df["status"])
    active = c["valid"] + c["invalid"]
    print(f"עובדים: {len(df):,}")
    for status in ("valid", "invalid", "no_base", "multi_period"):
        print(f"  {STATUS_HE[status]}: {c.get(status, 0):,}")
    if active:
        print(f"תקינות (פעילים): {c['valid'] / active * 100:.2f}%")
    print(f"נכתב: {xlsx_path}")
    print(f"נכתב: {md_path}")


if __name__ == "__main__":
    main_cli()
