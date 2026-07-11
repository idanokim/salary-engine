"""
extract_rules.py — build component_rules.json from the Progim engine workbook.

The חוקה (per-component computation rules) lives in the SACHAR sheet, block
AA3:DV11: row 9 holds the pay code, row 10 the name, row 7 the percentage rate
(resolved from `tosafot`), and row 11 the amount formula. Percentage components
reference an explicit list of OTHER columns in the same block — i.e., the exact
base composition (which pay codes the percentage applies to). This script
translates those column references back to pay codes so the salary engine can
recompute each percentage component from the slip's own recorded amounts:

    expected(code) = rate × Σ slip_amount(base codes)

Components whose Netunei Gimlai column H says 'ידני' (manual) have no formula —
the Progim engine takes the value straight from column J (manually reported).
The batch engine therefore accepts them as reported and never flags them.

Usage:
    python tools/extract_rules.py path/to/Progim.xlsm [component_rules.json]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

BLOCK_FIRST, BLOCK_LAST = "AA", "DV"
CODE_ROW, NAME_ROW, RATE_ROW, AMOUNT_ROW = 9, 10, 7, 11

# Codes whose Netunei Gimlai H column says 'ידני' — manually reported, the
# engine must accept the slip value as-is (Progim takes it from column J).
MANUAL_CODES = [4120, 4173, 4643, 4935]

# Multiple admissible rates for components whose rate depends on a per-worker
# attribute the גולמי file does not carry (role/certification/ministry wing),
# from the חוקה:
#   875  תוספת בתי משפט — BATEI MISHPAT 875 + the component note:
#         22 / 18.78 / 16.5 / 14.5 / 12.75% by ministry & track
#   858  אחוז ממשולב    — מוסמך 20% / לא מוסמך 15%
#   756  אחוז ממשולב    — per ministry: 104→30%, 109→7.5%, 111/180→15%, 170→8%
MULTI_RATE = {
    875: [0.22, 0.1878, 0.165, 0.145, 0.1275],
    858: [0.20, 0.15],
    756: [0.30, 0.15, 0.08, 0.075],
}

# Historical phase-in schedules (from `tos kibutzi`, `heskem 2016`, `MANMASH
# 2010`): the collective agreements ramped up over the years, so a slip from an
# earlier pay period legitimately carries an earlier official rate. A component
# matches if it equals ANY official rate × the correct base — wrong bases and
# arbitrary amounts still fail every admissible rate.
HISTORICAL_RATES = {
    # הסכם 2008/2009: פעימות מצטברות 1.2008 → 1.5%, 12.2008 → 3%, 12.2009 → 5%
    # (מח"ר קיבל 4.7% ישיר — 0.3% הוקצו לרווחה).
    4934: [0.015, 0.03, 0.047, 0.05],
    4994: [0.0225, 0.04, 0.0625, 0.0725],             # הסכם 2011: מדרגות 1.2011/1.2012/1.2013/7.2013
    5401: [0.01, 0.01375, 0.02125, 0.03, 0.03875],    # הסכם 2016: מדרגות 7.2016–12.2019
    5216: [0.05, 0.07, 0.08, 0.10],                   # מנמ"ש 2010: רמות × 8.2010/8.2011
    4624: [0.04, 0.036],                              # תוספת 99: מינהלי 4%, מח"ר 3.6% (tos kibutzi per droog-darga)
    # תוספת 2024 (הסכם המסגרת 2020–2027): פעימות מצטברות 12.2024 → 2%,
    # 4.2025 → 3.5% (מינהלי 3.94% per droog-darga), 4.2026 → 5%, 4.2027 → 6%.
    5533: [0.02, 0.035, 0.0394, 0.05, 0.06],
    741:  [0.30, 0.12],                               # בוררות מיסים: 30%; ותיקה 12% (מקבץ מדויק בנתוני 12.2011)
}

# Codes that belonged to a percentage's base in earlier pay periods and were
# dropped from the 2026 חוקה. Adding them is safe for current files (a worker
# without the code contributes 0) and removes hundreds of false flags on
# historical files — identified empirically: each miss's unaccounted base equals
# exactly that component's slip amount.
ERA_BASE_EXTRA = {
    4544: [1961, 1967, 5270, 636],                    # רפורמה ב' / ותק פעילות / מנמ"ש
    4934: [4147, 4651, 5270, 920, 1961, 1967, 636],   # תוספת רפורמה / תוספת שכר / ותק פעילות / 7.5% / רפורמה ב'
    4994: [4147, 4651, 5270, 920, 1961, 1967, 636],
    858:  [4624],
    4169: [4624],
}

# 920 (תוספת 7.5%) sums a fixed 257.37 into its base in the חוקה:
#   'tosafot'!BA2*(AA11+AC11+AD11+AE11+BI11+257.37)
BASE_CONSTANTS = {920: 257.37}

# The sample worker in the workbook has most eligibility switches off, so the
# resolved rate in SACHAR row 7 is 0 for components he doesn't get. The current
# rate still lives in the `tosafot` sheet's rate row — pull it from there.
# code -> (tosafot cell holding the rate)
RATE_FALLBACK_CELLS = {
    4169: "BD3",   # תוספת ענ"א 16.2%
    4932: "CA3",   # ענ"א 8%
    920:  "BA3",   # תוספת 7.5%
    4687: "BI3",   # תוספת שכר 22%
    697:  "Y3",    # תוספת מיוחדת
    659:  "W3",    # ת. קרימינולוג קליני
    798:  "AC3",   # דריכות בית חולים
    5216: "AH3",   # מנמ"ש 2010
    1053: "CF3",   # ענ"א
    4374: "CU3",   # נתיב ג
    4658: "CV3",   # גמול תפקיד
    5287: "CZ3",   # תוספת בט"פ
    5533: "DQ3",   # תוספת אחוזית 2024 (same 3.94% grid as 2029)
    2029: "DQ3",   # ר חדשה אחוזי
}


def parse_codes(cell_value):
    """A code cell is either an int or a string like '1037 / 4544' (aliases)."""
    if cell_value is None:
        return []
    if isinstance(cell_value, (int, float)):
        return [int(cell_value)]
    return [int(tok) for tok in re.findall(r"\d+", str(cell_value))]


def extract(xlsm_path: str) -> dict:
    wbf = openpyxl.load_workbook(xlsm_path, read_only=False, data_only=False)
    wbv = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    wsf, wsv = wbf["SACHAR"], wbv["SACHAR"]
    tosafot = {f"{get_column_letter(c)}{r}": v
               for r, row in enumerate(wbv["tosafot "].iter_rows(
                   min_row=1, max_row=6, values_only=True), start=1)
               for c, v in enumerate(row, start=1)}

    c0 = column_index_from_string(BLOCK_FIRST)
    c1 = column_index_from_string(BLOCK_LAST)

    # Map block column letter -> [codes], and read resolved rate values.
    sachar_vals = list(wsv.iter_rows(min_row=1, max_row=AMOUNT_ROW,
                                     min_col=c0, max_col=c1, values_only=True))
    col_codes, col_name, col_rate = {}, {}, {}
    for c in range(c0, c1 + 1):
        letter = get_column_letter(c)
        i = c - c0
        col_codes[letter] = parse_codes(sachar_vals[CODE_ROW - 1][i])
        col_name[letter] = sachar_vals[NAME_ROW - 1][i]
        rate = sachar_vals[RATE_ROW - 1][i]
        col_rate[letter] = float(rate) if isinstance(rate, (int, float)) and rate else None

    rules = {}
    for c in range(c0, c1 + 1):
        letter = get_column_letter(c)
        codes = col_codes[letter]
        if not codes or codes[0] in (99998, 99999):
            continue
        formula = wsf.cell(AMOUNT_ROW, c).value
        if not isinstance(formula, str):
            continue
        # A percentage rule looks like: (=... (<COL>11+<COL>11+...)*<RATE-REF>)
        # where RATE-REF is <this col>7 or a literal. Capture the base columns.
        base_cols = re.findall(r"([A-Z]{1,2})11", formula)
        base_cols = [b for b in base_cols if b != letter and column_index_from_string(b) >= c0]
        is_percent = bool(base_cols) and (f"{letter}7" in formula or "*" in formula)
        if not is_percent:
            continue
        base_codes = sorted({code for b in base_cols for code in col_codes.get(b, [])})
        if not base_codes:
            continue
        primary = max(codes)  # the code seen in modern payroll dumps
        rates = MULTI_RATE.get(primary)
        if rates is None:
            rate = col_rate[letter]
            if rate is None and primary in RATE_FALLBACK_CELLS:
                v = tosafot.get(RATE_FALLBACK_CELLS[primary])
                rate = float(v) if isinstance(v, (int, float)) and v else None
            rates = sorted(set(([rate] if rate is not None else [])
                               + HISTORICAL_RATES.get(primary, [])))
            if not rates:
                continue  # rate not resolvable — skip (never guess a rule)
        # The Pension Authority dumps split שכר משולב (10002) into
        # יסוד משולב (1) + תוספת ותק (2); wherever the base includes the
        # combined code, the split pair is the same amount.
        base_codes = sorted(set(base_codes) | set(ERA_BASE_EXTRA.get(primary, [])))
        if 10002 in base_codes:
            base_codes = sorted(set(base_codes) | {1, 2})
        rule = {
            "codes": codes,
            "name": str(col_name[letter] or "").strip(),
            "type": "percent",
            "rates": rates,
            "base_codes": base_codes,
        }
        if primary in BASE_CONSTANTS:
            rule["base_const"] = BASE_CONSTANTS[primary]  # × job% at runtime
        rules[primary] = rule

    for code in MANUAL_CODES:
        rules[code] = {
            "codes": [code], "type": "manual",
            "name": {4120: "השלמת שכר", 4173: "תוספת אישית",
                     4643: "השלמת שכר", 4935: "תוספת אמון"}[code],
            "note": "ידני — הערך מדווח (Netunei Gimlai עמודה J), אין נוסחה לאימות",
        }
    wbv.close()
    return rules


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else "component_rules.json"
    rules = extract(src)
    Path(dst).write_text(json.dumps(rules, ensure_ascii=False, indent=1), encoding="utf-8")
    n_pct = sum(1 for r in rules.values() if r["type"] == "percent")
    n_man = sum(1 for r in rules.values() if r["type"] == "manual")
    print(f"Wrote {dst}: {n_pct} percent rules, {n_man} manual codes")


if __name__ == "__main__":
    main()
