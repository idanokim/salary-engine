# -*- coding: utf-8 -*-
"""
unified_report.py — run the engine on one or more monthly גולמי files and write
a polished, workable unified workbook:

  לוח בקרה     — KPI tiles + per-file summary with conditional formatting
  שגויים לבדיקה — the work queue: only invalid slips, sorted by |gap|, filterable
  פר עובד      — every employee from every file as a filterable Excel table
  פילוח משרדים — ministry × validity aggregated across all files

Usage:
    python tools/unified_report.py file1.xlsx [file2.xlsx ...] --out unified.xlsx
"""

import argparse
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import main as engine

# ---- palette (light surface; status colors paired with a text label, never alone)
NAVY = "FF1E3A5F"          # brand header
NAVY_TEXT = "FFFFFFFF"
TILE_BG = "FFF4F5F7"
GOOD_TXT, GOOD_BG = "FF0B7A0B", "FFE7F4E7"
BAD_TXT, BAD_BG = "FFA82626", "FFFBE9E9"
WARN_TXT, WARN_BG = "FF8A5A00", "FFFEF3D8"
MUTED = "FF6B7280"
BAR_BLUE = "FF2A78D6"
BORDER = Side(style="thin", color="FFD9DDE3")
THIN_BOX = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

MONEY = "#,##0.00"
INT = "#,##0"
PCT1 = '0.00"%"'

STATUS_HE = {"valid": "תקין", "invalid": "שגוי",
             "no_base": "ללא שכר בסיס פעיל", "multi_period": "רטרו / רב-תקופתי"}


def _header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, color=NAVY_TEXT, size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 24


def _kpi(ws, row, col, span, title, value, color=None, fmt=None):
    """A merged stat tile: muted title above a large value."""
    c0, c1 = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{c0}{row}:{c1}{row}")
    ws.merge_cells(f"{c0}{row + 1}:{c1}{row + 1}")
    t = ws.cell(row=row, column=col, value=title)
    t.font = Font(size=10, color=MUTED)
    t.alignment = Alignment(horizontal="center", vertical="bottom")
    v = ws.cell(row=row + 1, column=col, value=value)
    v.font = Font(size=16, bold=True, color=color or "FF0B0B0B")
    v.alignment = Alignment(horizontal="center", vertical="top")
    if fmt:
        v.number_format = fmt
    for r in (row, row + 1):
        for cc in range(col, col + span):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TILE_BG)


def pay_month_of(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    month = None
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        for v in row:
            if isinstance(v, datetime):
                month = v
                break
        if month:
            break
    wb.close()
    return month


def collect(paths):
    """Run the engine per file; return (summary rows, per-employee rows)."""
    lookups = engine.get_lookups()
    files = sorted(paths, key=lambda p: (pay_month_of(p) or datetime(2099, 1, 1),
                                         Path(p).name))
    summary, per_emp = [], []
    for path in files:
        d = pay_month_of(path)
        month = d.strftime("%m/%Y") if d else Path(path).stem[:12]
        short = Path(path).stem.split("-", 1)[-1][:20] or Path(path).stem[:20]
        workers = engine.load_golmi(path)
        entries = engine.run_engine_full(workers, lookups)
        c = Counter(e["result"].status for e in entries)
        active = c["valid"] + c["invalid"]
        summary.append({
            "month": month, "file": short, "workers": len(workers),
            "valid": c["valid"], "invalid": c["invalid"],
            "no_base": c["no_base"], "multi": c["multi_period"],
            "acc": round(c["valid"] / active * 100, 2) if active else 0.0,
        })
        print(f"  {month}  {len(workers):>7,} עובדים · שגויים {c['invalid']:>5,} "
              f"· {summary[-1]['acc']:.2f}%  ({short})")
        for e in entries:
            r, flags = e["result"], e["comp_flags"]
            bs = sum((cp.expected or 0.0) for cp in r.components
                     if cp.code in engine.BASE_CODES)
            bc = sum(cp.amount for cp in r.components if cp.calculated)
            # gap in גמולי השתלמות (slip minus standard), from the gmul flags
            gmul_diff = round(sum(chk["slip"] - chk["expected"]
                                  for code, chk in flags.items()
                                  if code in (667, 897)), 2) or None
            per_emp.append({
                "month": month, "file": short, "worker_id": r.worker_id,
                "ministry": r.ministry_name, "darga": r.darga_label,
                "vatek": r.vatek_calculated, "job_pct": r.job_pct,
                "base_slip": round(bs, 2),
                "base_calc": round(bc, 2) if bc else None,
                "base_diff": round(bc - bs, 2) if bc else None,
                "gmul_diff": gmul_diff,
                "total_slip": r.expected_total, "total_calc": r.total,
                "total_diff": r.total_diff, "status": r.status,
                "flags": "; ".join(
                    f"{k} ({v['name']}): {v['slip']} במקום {v['expected']}"
                    for k, v in sorted(flags.items())),
                "diag": "; ".join(r.errors),
            })
    return summary, per_emp


EMP_COLS = [
    ("month", "חודש שכר", 11, None), ("file", "קובץ", 18, None),
    ("worker_id", "מסד עובד", 12, INT), ("ministry", "משרד", 22, None),
    ("darga", "דרגה", 8, None), ("vatek", "ותק", 8, None),
    ("job_pct", "חלקיות", 8, None), ("base_slip", "בסיס בתלוש", 13, MONEY),
    ("base_calc", "בסיס מחושב", 13, MONEY), ("base_diff", "הפרש בסיס", 12, MONEY),
    ("gmul_diff", "הפרש גמולים", 13, MONEY),
    ("total_slip", "סכום בתלוש", 13, MONEY), ("total_calc", "סכום מחושב", 13, MONEY),
    ("total_diff", "הפרש כולל", 12, MONEY), ("status_he", "סטטוס", 16, None),
    ("flags", "רכיבים חריגים", 30, None), ("diag", "אבחון", 30, None),
]


def _emp_sheet(wb, title, rows, table_name, highlight_invalid):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    _header_row(ws, 1, [he for _, he, _, _ in EMP_COLS],
                [w for _, _, w, _ in EMP_COLS])
    inv_font = Font(color=BAD_TXT, bold=True)
    inv_fill = PatternFill("solid", fgColor=BAD_BG)
    ok_font = Font(color=GOOD_TXT)
    warn_font = Font(color=WARN_TXT)
    for r_i, row in enumerate(rows, start=2):
        vals = [row.get(k) if k != "status_he" else STATUS_HE[row["status"]]
                for k, _, _, _ in EMP_COLS]
        for c_i, ((key, _, _, fmt), v) in enumerate(zip(EMP_COLS, vals), start=1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            if fmt:
                cell.number_format = fmt
            if key == "status_he":
                st = row["status"]
                cell.font = (inv_font if st == "invalid"
                             else ok_font if st == "valid" else warn_font)
                if st == "invalid":
                    cell.fill = inv_fill
            elif key in ("base_diff", "gmul_diff", "total_diff") and row["status"] == "invalid":
                cell.font = inv_font
        if highlight_invalid and row["status"] == "invalid":
            ws.cell(row=r_i, column=1).fill = inv_fill
    if rows:
        ref = f"A1:{get_column_letter(len(EMP_COLS))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight15",
                                              showRowStripes=True)
        ws.add_table(table)
    return ws


def write_workbook(summary, per_emp, out_path):
    wb = openpyxl.Workbook()

    # ---- לוח בקרה ------------------------------------------------------------
    ws = wb.active
    ws.title = "לוח בקרה"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "בדיקת התאמת תלושים — דוח מאוחד"
    t.font = Font(size=16, bold=True, color=NAVY)
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"{len(summary)} קבצים · הופק "
               f"{datetime.now().strftime('%d/%m/%Y %H:%M')} · "
               "תלוש תקין = הפרש עד ₪1 מהחישוב לפי טבלאות השכר")
    s.font = Font(size=10, color=MUTED)

    tot = Counter()
    for r in summary:
        for k in ("workers", "valid", "invalid", "no_base", "multi"):
            tot[k] += r[k]
    active = tot["valid"] + tot["invalid"]
    acc = round(tot["valid"] / active * 100, 2) if active else 0.0
    _kpi(ws, 4, 1, 1, "סה\"כ עובדים", tot["workers"], fmt=INT)
    _kpi(ws, 4, 2, 1, "תקינים", tot["valid"], GOOD_TXT, INT)
    _kpi(ws, 4, 3, 1, "שגויים — לבדיקה", tot["invalid"], BAD_TXT, INT)
    _kpi(ws, 4, 4, 1, "% תקינות (פעילים)", acc / 100, GOOD_TXT if acc >= 99 else WARN_TXT, "0.00%")
    _kpi(ws, 4, 5, 1, "ללא בסיס פעיל", tot["no_base"], WARN_TXT, INT)
    _kpi(ws, 4, 6, 1, "רטרו / רב-תקופתי", tot["multi"], WARN_TXT, INT)

    head_r = 7
    labels = ["חודש שכר", "קובץ", "עובדים", "תקין", "שגוי",
              "ללא בסיס", "רטרו", "% תקינות"]
    _header_row(ws, head_r, labels, [11, 20, 11, 11, 9, 10, 8, 11])
    for i, r in enumerate(summary, start=head_r + 1):
        vals = [r["month"], r["file"], r["workers"], r["valid"], r["invalid"],
                r["no_base"], r["multi"], r["acc"] / 100]
        for c_i, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c_i, value=v)
            cell.border = THIN_BOX
            if c_i in (3, 4, 5, 6, 7):
                cell.number_format = INT
            if c_i == 4:
                cell.font = Font(color=GOOD_TXT)
            if c_i == 5 and r["invalid"]:
                cell.font = Font(color=BAD_TXT, bold=True)
            if c_i == 8:
                cell.number_format = "0.00%"
    last = head_r + len(summary)
    trow = last + 1
    tvals = ["סה\"כ", "", tot["workers"], tot["valid"], tot["invalid"],
             tot["no_base"], tot["multi"], acc / 100]
    for c_i, v in enumerate(tvals, start=1):
        cell = ws.cell(row=trow, column=c_i, value=v)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="double", color=NAVY))
        if c_i in (3, 4, 5, 6, 7):
            cell.number_format = INT
        if c_i == 8:
            cell.number_format = "0.00%"
    rng = f"H{head_r + 1}:H{last}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["0.99"],
        font=Font(color=GOOD_TXT), fill=PatternFill("solid", fgColor=GOOD_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["0.95", "0.9899"],
        font=Font(color=WARN_TXT), fill=PatternFill("solid", fgColor=WARN_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0.95"],
        font=Font(color=BAD_TXT), fill=PatternFill("solid", fgColor=BAD_BG)))
    ws.conditional_formatting.add(
        f"C{head_r + 1}:C{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=BAR_BLUE, showValue=True))
    ws.conditional_formatting.add(
        f"E{head_r + 1}:E{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="FFD03B3B", showValue=True))

    # ---- שגויים לבדיקה --------------------------------------------------------
    inv = [r for r in per_emp if r["status"] == "invalid"]
    inv.sort(key=lambda r: abs(r["base_diff"] if r["base_diff"] is not None
                               else (r["total_diff"] or 0)), reverse=True)
    _emp_sheet(wb, "שגויים לבדיקה", inv, "Invalids", highlight_invalid=False)

    # ---- פר עובד ---------------------------------------------------------------
    _emp_sheet(wb, "פר עובד", per_emp, "PerEmployee", highlight_invalid=True)

    # ---- פילוח משרדים ----------------------------------------------------------
    agg = defaultdict(lambda: Counter())
    for r in per_emp:
        agg[r["ministry"] or "—"][r["status"]] += 1
    ws4 = wb.create_sheet("פילוח משרדים")
    ws4.sheet_view.rightToLeft = True
    ws4.freeze_panes = "A2"
    _header_row(ws4, 1, ["משרד / גוף", "עובדים", "תקין", "שגוי", "% תקינות (פעילים)"],
                [26, 11, 11, 9, 16])
    rows = sorted(agg.items(), key=lambda kv: -sum(kv[1].values()))
    for i, (name, c) in enumerate(rows, start=2):
        act = c["valid"] + c["invalid"]
        vals = [name, sum(c.values()), c["valid"], c["invalid"],
                (c["valid"] / act) if act else None]
        for c_i, v in enumerate(vals, start=1):
            cell = ws4.cell(row=i, column=c_i, value=v)
            cell.border = THIN_BOX
            if c_i in (2, 3, 4):
                cell.number_format = INT
            if c_i == 4 and c["invalid"]:
                cell.font = Font(color=BAD_TXT, bold=True)
            if c_i == 5:
                cell.number_format = "0.0%"
    last4 = len(rows) + 1
    ws4.conditional_formatting.add(
        f"B2:B{last4}", DataBarRule(start_type="num", start_value=0,
                                    end_type="max", color=BAR_BLUE, showValue=True))
    rng4 = f"E2:E{last4}"
    ws4.conditional_formatting.add(rng4, CellIsRule(
        operator="greaterThanOrEqual", formula=["0.99"],
        fill=PatternFill("solid", fgColor=GOOD_BG)))
    ws4.conditional_formatting.add(rng4, CellIsRule(
        operator="lessThan", formula=["0.95"],
        fill=PatternFill("solid", fgColor=BAD_BG)))
    tbl4 = Table(displayName="Ministries", ref=f"A1:E{last4}")
    tbl4.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
    ws4.add_table(tbl4)

    # ---- מגמה בין-חודשית: native Excel charts on the dashboard -----------------
    months = sorted({r["month"] for r in summary},
                    key=lambda m: (m.split("/")[-1], m.split("/")[0]))
    if len(months) > 1:
        per_month = {m: Counter() for m in months}
        for r in summary:
            for k in ("workers", "valid", "invalid"):
                per_month[r["month"]][k] += r[k]
        anchor = trow + 3
        ws.cell(row=anchor - 1, column=1, value="מגמה בין-חודשית").font = \
            Font(bold=True, color=NAVY)
        _header_row(ws, anchor, ["חודש", "% תקינות", "שגויים"], [11, 11, 9])
        for i, m in enumerate(months, start=anchor + 1):
            c = per_month[m]
            act = c["valid"] + c["invalid"]
            ws.cell(row=i, column=1, value=m).border = THIN_BOX
            pc = ws.cell(row=i, column=2, value=(c["valid"] / act) if act else None)
            pc.number_format = "0.00%"
            pc.border = THIN_BOX
            ic = ws.cell(row=i, column=3, value=c["invalid"])
            ic.number_format = INT
            ic.border = THIN_BOX
        from openpyxl.chart import BarChart, LineChart, Reference
        lastm = anchor + len(months)
        line = LineChart()
        line.title = "% תקינות לפי חודש"
        line.height, line.width = 7, 13
        line.y_axis.numFmt = "0%"
        line.add_data(Reference(ws, min_col=2, min_row=anchor, max_row=lastm),
                      titles_from_data=True)
        line.set_categories(Reference(ws, min_col=1, min_row=anchor + 1, max_row=lastm))
        ws.add_chart(line, f"E{anchor - 1}")
        bar = BarChart()
        bar.title = "שגויים לפי חודש"
        bar.height, bar.width = 7, 13
        bar.add_data(Reference(ws, min_col=3, min_row=anchor, max_row=lastm),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=anchor + 1, max_row=lastm))
        ws.add_chart(bar, f"E{anchor + 15}")

        # ---- שינויים בין חודשים: עובד תקין שהפך שגוי ----------------------------
        month_idx = {m: i for i, m in enumerate(months)}
        by_worker = defaultdict(dict)
        for r in per_emp:
            if r["status"] in ("valid", "invalid"):
                by_worker[r["worker_id"]][r["month"]] = r
        regressions = []
        for wid, per_m in by_worker.items():
            ms = sorted(per_m, key=lambda m: month_idx.get(m, 99))
            for a, b in zip(ms, ms[1:]):
                if per_m[a]["status"] == "valid" and per_m[b]["status"] == "invalid":
                    cur = per_m[b]
                    regressions.append({
                        "worker_id": wid, "ministry": cur["ministry"],
                        "darga": cur["darga"], "from_m": a, "to_m": b,
                        "total_diff": cur["total_diff"], "flags": cur["flags"],
                        "diag": cur["diag"],
                    })
        ws5 = wb.create_sheet("שינויים בין חודשים")
        ws5.sheet_view.rightToLeft = True
        ws5.freeze_panes = "A3"
        ws5.merge_cells("A1:H1")
        h = ws5["A1"]
        h.value = ("עובדים שהיו תקינים בחודש מוקדם והפכו שגויים בחודש מאוחר — "
                   f"{len(regressions)} מקרים · ממוין לפי גודל הפער")
        h.font = Font(bold=True, color=NAVY)
        _header_row(ws5, 2, ["מסד עובד", "משרד", "דרגה", "תקין בחודש",
                             "שגוי בחודש", "הפרש כולל", "רכיבים חריגים", "אבחון"],
                    [12, 22, 8, 12, 12, 12, 34, 34])
        regressions.sort(key=lambda r: abs(r["total_diff"] or 0), reverse=True)
        for i, r in enumerate(regressions, start=3):
            vals = [r["worker_id"], r["ministry"], r["darga"], r["from_m"],
                    r["to_m"], r["total_diff"], r["flags"], r["diag"]]
            for c_i, v in enumerate(vals, start=1):
                cell = ws5.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
                if c_i == 1:
                    cell.number_format = INT
                if c_i == 6:
                    cell.number_format = MONEY
                    cell.font = Font(color=BAD_TXT, bold=True)

    wb.save(out_path)


def main_cli():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("files", nargs="+", help="גולמי .xlsx files")
    ap.add_argument("--out", default="unified.xlsx")
    args = ap.parse_args()
    t0 = time.time()
    summary, per_emp = collect(args.files)
    print(f"עיבוד: {time.time() - t0:.0f}ש · כותב workbook ({len(per_emp):,} שורות)...")
    write_workbook(summary, per_emp, args.out)
    inv = sum(1 for r in per_emp if r["status"] == "invalid")
    print(f"נכתב: {args.out} · {len(per_emp):,} רשומות · {inv:,} שגויים")


if __name__ == "__main__":
    main_cli()
